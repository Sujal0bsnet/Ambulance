from tkinter import*
from tkinter import ttk
from tkinter import Tk, Label,PhotoImage
from PIL import ImageTk,Image
from tkinter import messagebox
import sqlite3
import ast
import database
import firstaid
import cutomersupport
import user
import pathlib
import os
import openpyxl,xlrd
from openpyxl import Workbook,load_workbook
from tkinter .ttk import Combobox
import pandas as pd


root=Tk()
root.title("Login")
root.geometry("925x500+300+200")
root.configure(bg="#fff")
root.resizable(0,0)
root.iconbitmap("a.ico")
database.createDatabaseIfNotExists()

abc=ImageTk.PhotoImage(Image.open("b.png"))
Label(root,image=abc,bg="white").place(x=50,y=50)

  

################################Greeting###############################################
# screen=None
def signin():
    username = name.get()
    password = system.get()
    try:
        database.loginUser(username,password)
        root.destroy()
        global screen
        screen = Tk()
        screen.title("life Booking System")
        screen.geometry("925x500+300+200")
        screen.iconbitmap("intro2.ico")
        screen.configure(bg="red")

        abc=ImageTk.PhotoImage(Image.open("00.png"))
        Label(screen,image=abc,bg="red").place(x=115,y=-50)

        welcome_message = "Hello {}, welcome to the Ambulance Booking System!".format(username)
        Label(screen, text=welcome_message, bg="white", font=("Garamond", 12, "bold")).place(x=250,y=250)
        btn=Button(screen,text="Enter",fg="white",bg="black",font=("Times New Roman",10,'bold'),command=dashboard)
        btn.place(x=430,y=275)

        screen.mainloop()
    except BaseException as ex:
        messagebox.showerror("Error",str(ex))  
  #####################################Dashboard##########################################################################
    
def dashboard():
    screen.destroy()
    global root2
    root2=Tk()
    root2.title('Save Life')
    root2.geometry('1720x765')
    root2.iconbitmap('oo.ico')
    frame=Frame(root2,height=30,width=1500,bg='red')
    frame.place(x=10,y=5)

    topic=Label(frame,text="Always Ready To Help, Anytime & Anywhere",fg="white",bg="red",font=("Helvetica"))
    topic.place(x=545,y=5)

    frame0=Frame(root2,height=475,width=1475,bg='white')
    frame0.place(x=10,y=75)

    frame00=Frame(frame0,height=445,width=340,bg='')
    frame00.place(x=20,y=35)

    img0=ImageTk.PhotoImage(Image.open("abc.png"))
    Label(frame,image=img0, bg="red").place(x=-2,y=-2)

    img01=ImageTk.PhotoImage(Image.open("01.png"))
    Label(frame,image=img01,height=30,bg="red").place(x=1475,y=0)

    frame1=Frame(root2,height=5,width=1050,bg='red')
    frame1.place(x=210,y=325)

    frame2=Frame(root2,height=475,width=5,bg='red')
    frame2.place(x=725,y=75)

    frame3=Frame(root2,height=160,width=1500,bg='red')
    frame3.place(x=10,y=575)


    img=ImageTk.PhotoImage(Image.open("15.png"))
    Label(frame0,image=img,height=96,width=96,bg="white").place(x=465,y=55)

    img2=ImageTk.PhotoImage(Image.open("16.png"))
    Label(frame0,image=img2,height=96,width=96,bg="white").place(x=925,y=55)

    img3=ImageTk.PhotoImage(Image.open("171.png"))
    Label(frame0,image=img3,height=96,width=96,bg="white").place(x=465,y=285)

    img4=ImageTk.PhotoImage(Image.open("18.png"))
    Label(frame0,image=img4,height=96,width=96,bg="white").place(x=925,y=285)

    button_l1=Button(frame0,text='Find Ambulance',fg='white',bg='red',font=('arial',12,'bold'),command=Find_Ambulance)
    button_l1.place(x=450,y=145)

    button_l2=Button(frame0,text='Find Hospitals',fg='white',bg='red',font=('arial',12,'bold'),command=Find_Hospital)
    button_l2.place(x=920,y=145)

    button_l3=Button(frame0,text='Find FIRST AIDS',fg='white',bg='red',font=('arial',12,'bold'),command=first_aid)
    button_l3.place(x=450,y=380)

    button_l4=Button(frame0,text='Customer Care',fg='white',bg='red',font=('arial',12,'bold'),command=customercare)
    button_l4.place(x=915,y=385)
    location=ImageTk.PhotoImage(Image.open("location.png"))
    Label(frame3,image=location,height=25,bg="red").place(x=1015,y=20)

    f4=Label(frame3,text="Head Office, Dilli Bazar,Kathmandu 44600,Nepal",fg='black',bg='red',font=('arial',12,'bold'))
    f4.place(x=1075,y=20)

    imgf1=ImageTk.PhotoImage(Image.open("f1.png"))
    Label(frame3,image=imgf1,height=25,bg="red").place(x=1015,y=60)

    f5=Label(frame3,text="Contact Us:- +977-9864220424",fg='black',bg='red',font=('arial',12,'bold'))
    f5.place(x=1075,y=60)

    mail=ImageTk.PhotoImage(Image.open("mail.png"))
    Label(frame3,image=mail,height=25,bg="red").place(x=1015,y=90)

    f6=Label(frame3,text="info@ambulancebook.com",fg='black',bg='red',font=('arial',12,'bold'))
    f6.place(x=1075,y=90)

    f7=Label(frame3,text="Designed And Developed By Batch34(B).",bg='red',fg='violet',font=('arial',12,'bold'))
    f7.place(x=570,y=125)

   



    root2.mainloop()

#############################Find Hospital######################################
# image_var = None 
def Find_Hospital():
    
    # global image_var
    root3=Toplevel()
    
    root3.title("Find Hospital")
    root3.geometry("925x500+300+200")
    root3.iconbitmap("c.ico")
    root3.configure(bg="Red")

    root3.resizable(0,0)

    frame=Frame(root3,width=345,height=50,bg="#fff")

    frame.place(x=280,y=10)

    topic=Label(frame,text="Find Hospital Near You",fg="red",bg="white",font=("Times New roamn",19,"bold"))
    topic.place(x=30,y=10)

################Om hospital################################################
    def appointment():
            root3.destroy()
            root4=Tk()
            root4.title("Om Appointment Booking System",)
            root4.geometry("1000x500")
            root4.iconbitmap("88.ico")
            root4.configure(bg="Gray")

            WIN=ttk.Treeview(root4)
            Hospital="Om Hospital"

            def reverse(tuples):
                new=tuples[::-1]
                return new

            def insert(id,name,age,address,number,date,time):
                database.addAppointment(name,age,address,number,date,time,1,user.username)
                messagebox.showinfo("omhospital","Booked sucessfully")
    

            def delete_data_database(data):
                database.deleteAppointment(int(data))
                messagebox.showinfo("omhospital","Deleted sucessfully")

            def update_database(id,name,age,address,number,date,time):
                database.updateAppointment(int(id),name,age,address,number,date,time)
                messagebox.showinfo("omhospital","Updated sucessfully")

            def read():
                
                if(not user.isAdmin):
                    return database.getAppointmentListForUser(1,user.username)
                else:
                   return database.getAppointmentListForAdmin(1)

            def insert_data():

#   id=str(entryID.get())
                  name=str(entryName.get())
                  age=str(entryAge.get())
                  address=str(entryAddress.get())
                  number=str(entryNumber.get())
                  date=str(entryDate.get())
                  time=str(entryTime.get())

                  if id==""or id==" ":
                     print("Error Inserting Id")
                  if name==""or name==" ":
                     print("Error Inserting Name") 
                  if age==""or age==" ":
                     print("Error Inserting Age") 
                  if address==""or address==" ":
                     print("Error Inserting Address")
                  if number==""or number==" ":
                     print("Error Inserting Number") 
                  if date==""or date==" ":
                     print("Error Inserting Date") 
                  if time==""or time==" ":
                     print("Error Inserting Time")

                  else:
                      insert(str(id),str(name),str(age),str(address),str(number),str(date),str(time))

                  for data in WIN.get_children():
                      WIN.delete(data)
                  for result in reverse(read()):
                      WIN.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))


    
                  WIN.tag_configure("orow",background="Red",font=("Times New Roman",8))
                  WIN.grid(row=1,column=4,columnspan=4,rowspan=5,padx=10,pady=10)
 

            def delete_data():
                selected_app=WIN.selection()[0]
                deleteData=str(WIN.item(selected_app)["values"][0])
                delete_data_database(deleteData)

                for data in WIN.get_children():
                      WIN.delete(data)
                for result in reverse(read()):
                      WIN.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
    
                WIN.tag_configure("orow",background="Red",font=("Times New Roman",8))
                WIN.grid(row=1,column=4,columnspan=4,rowspan=5,padx=10,pady=10) 
      
            def update_data():             
                if(len(WIN.selection())<1): return
                selected_app=WIN.selection()[0]
                updatedata=str(WIN.item(selected_app)["values"][0])
                update_database(updatedata,entryName.get(),entryAge.get(),entryAddress.get(),entryNumber.get(),entryDate.get(),entryTime.get())

                for data in WIN.get_children():
                  WIN.delete(data)
                for result in reverse(read()):
                  WIN.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
    
                WIN.tag_configure("orow",background="Red",font=("Times New Roman",8))
                WIN.grid(row=1,column=4,columnspan=4,rowspan=5,padx=10,pady=10)

    


            title=Label(root4,text="Om Appointment",font=("Times New Roman",30),bd=2)
            title.grid(row=0,column=0,columnspan=8,padx=20,pady=20)

# id=Label(root,text="ID",font=("Times New Roman",15))
            name=Label(root4,text="Name",font=("Times New Roman",17),bg="gray",fg="white")
            age=Label(root4,text="Age",font=("Times New Roman",17),bg="gray",fg="white")
            address=Label(root4,text="Address",font=("Times New Roman",17),bg="gray",fg="white")
            number=Label(root4,text="Phone Number",font=("Times New Roman",17),bg="gray",fg="white")
            date=Label(root4,text="Date",font=("Times New Roman",17),bg="gray",fg="white")
            time=Label(root4,text="Time",font=("Times New Roman",17),bg="gray",fg="white")

# id.grid(row=1,column=0,padx=5,pady=5)
            name.grid(row=2,column=0,padx=5,pady=5)
            age.grid(row=3,column=0,padx=5,pady=5)
            address.grid(row=4,column=0,padx=5,pady=5)
            number.grid(row=5,column=0,padx=5,pady=5)
            date.grid(row=6,column=0,padx=5,pady=5)
            time.grid(row=7,column=0,padx=5,pady=5)

# entryID=Entry(root,width=25,bd=5,font=("Times New Roman",15))
            entryName=Entry(root4,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
            entryAge=Entry(root4,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
            entryAddress=Entry(root4,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
            entryNumber=Entry(root4,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
            entryDate=Entry(root4,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
            entryTime=Entry(root4,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)

# entryID.grid(row=1,column=1,columnspan=3,padx=5,pady=5)
            entryName.grid(row=2,column=1,columnspan=3,padx=5,pady=5)
            entryAge.grid(row=3,column=1,columnspan=3,padx=5,pady=5)
            entryAddress.grid(row=4,column=1,columnspan=3,padx=5,pady=5)
            entryNumber.grid(row=5,column=1,columnspan=3,padx=5,pady=5)
            entryDate.grid(row=6,column=1,columnspan=3,padx=5,pady=5)
            entryTime.grid(row=7,column=1,columnspan=3,padx=5,pady=5)

            submit=Button(root4,text="Submit",padx=1,pady=1,width=5,bd=1,font=("Times New Roman",15),bg="RED",fg="white",command=insert_data)
            submit.grid(row=8,column=1)

            update=Button(root4,text="Update",padx=1,pady=1,width=5,bd=1,font=("Times New Roman",15),bg="Red",fg="white",command=update_data)
            update.grid(row=8,column=2)

            delete=Button(root4,text="Delete",padx=1,pady=1,width=5,bd=1,font=("Times New Roman",15),bg="Red",fg="white",command=delete_data)
            delete.grid(row=8,column=3)

#####for treeview
            style=ttk.Style()
            style.configure("Treeview.Heading",font=("Times New Roman",15))


            WIN["columns"]=("ID","Name","Age","Address","Number","Date","Time")
            WIN.column("#0", width=0,stretch=NO)
            WIN.column("ID",anchor=W,width=30)
            WIN.column("Name",anchor=W,width=100)
            WIN.column("Age",anchor=W,width=50)
            WIN.column("Address",anchor=W,width=90)
            WIN.column("Number",anchor=W,width=90)
            WIN.column("Date",anchor=W,width=60)
            WIN.column("Time",anchor=W,width=60)

# WIN.heading("#0", width=0,stretch=NO)
            WIN.heading("ID",text="ID",anchor=W)
            WIN.heading("Name",text="Name",anchor=W)
            WIN.heading("Age",text="Age",anchor=W)
            WIN.heading("Address",text="Address",anchor=W)
            WIN.heading("Number",text="Number",anchor=W)
            WIN.heading("Date",text="Date",anchor=W)
            WIN.heading("Time",text="Time",anchor=W)

            for data in WIN.get_children():
                WIN.delete(data)
            for result in reverse(read()):
                WIN.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
    
    
            WIN.tag_configure("orow",background="Red",font=("Times New Roman",8))
            WIN.grid(row=1,column=4,columnspan=4,rowspan=10,padx=10,pady=10) 


            root4.mainloop()


##################manmohan hospital##########
    def booking():
        root3.destroy()
        global root5  
        root5=Tk()
        root5.title("Manmohan Appointment Booking System",)
        root5.geometry("1000x500")
        root5.iconbitmap("88.ico")
        root5.configure(bg="Gray")

        win=ttk.Treeview(root5)
        Hospital="Manmohan Hospital"

        def reverse(tuples):
            new=tuples[::-1]
            return new

        def insert(id,name,age,address,number,date,time):
            
                database.addAppointment(name,age,address,number,date,time,3,user.username)
                messagebox.showinfo("Manmohan Hospital","Booked sucessfully")
            
        def delete_data_database(data):
            database.deleteAppointment(int(data))
            messagebox.showinfo("Manmohan Hospital","Deleted sucessfully")
            
           

        def update_database(id,name,age,address,number,date,time):
            database.updateAppointment(int(id),name,age,address,number,date,time)
            messagebox.showinfo("Manmohan Hospital","Updated sucessfully")
            

        def read():
                if(not user.isAdmin):
                    return database.getAppointmentListForUser(3,user.username)
                else:
                   return database.getAppointmentListForAdmin(3)

        def insert_data():
 
#            id=str(entryID.get())
             name=str(entryName.get())
             age=str(entryAge.get())
             address=str(entryAddress.get())
             number=str(entryNumber.get())
             date=str(entryDate.get())
             time=str(entryTime.get())

             if id==""or id==" ":
              print("Error Inserting Id")
             if name==""or name==" ":
              print("Error Inserting Name") 
             if age==""or age==" ":
              print("Error Inserting Age") 
             if address==""or address==" ":
              print("Error Inserting Address")
             if number==""or number==" ":
              print("Error Inserting Number") 
             if date==""or date==" ":
              print("Error Inserting Date") 
             if time==""or time==" ":
              print("Error Inserting Time")

             else:
              insert(str(id),str(name),str(age),str(address),str(number),str(date),str(time))

             for data in win.get_children():
              win.delete(data)
             for result in reverse(read()):
              win.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))


    
             win.tag_configure("orow",background="Red",font=("Times New Roman",8))
             win.grid(row=1,column=4,columnspan=4,rowspan=5,padx=10,pady=10)
 

        def delete_data():
              selected_app=win.selection()[0]
              deleteData=str(win.item(selected_app)["values"][0])
              delete_data_database(deleteData)

              for data in win.get_children():
                 win.delete(data)
              for result in reverse(read()):
                 win.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
    
                 win.tag_configure("orow",background="Red",font=("Times New Roman",8))
                 win.grid(row=1,column=4,columnspan=4,rowspan=5,padx=10,pady=10) 
      
        def update_data():             
             if(len(win.selection())<1): return
             selected_app=win.selection()[0]
             updatedata=str(win.item(selected_app)["values"][0])
             update_database(updatedata,entryName.get(),entryAge.get(),entryAddress.get(),entryNumber.get(),entryDate.get(),entryTime.get())

             for data in win.get_children():
                 win.delete(data)
             for result in reverse(read()):
                 win.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
    
             win.tag_configure("orow",background="Red",font=("Times New Roman",8))
             win.grid(row=1,column=4,columnspan=4,rowspan=5,padx=10,pady=10)

    


        title=Label(root5,text="Manmohan Appointment",font=("Times New Roman",30),bd=2)
        title.grid(row=0,column=0,columnspan=8,padx=20,pady=20)

# id=Label(root,text="ID",font=("Times New Roman",15))
        name=Label(root5,text="Name",font=("Times New Roman",17),bg="gray",fg="white")
        age=Label(root5,text="Age",font=("Times New Roman",17),bg="gray",fg="white")
        address=Label(root5,text="Address",font=("Times New Roman",17),bg="gray",fg="white")
        number=Label(root5,text="Phone Number",font=("Times New Roman",17),bg="gray",fg="white")
        date=Label(root5,text="Date",font=("Times New Roman",17),bg="gray",fg="white")
        time=Label(root5,text="Time",font=("Times New Roman",17),bg="gray",fg="white")

# id.grid(row=1,column=0,padx=5,pady=5)
        name.grid(row=2,column=0,padx=5,pady=5)
        age.grid(row=3,column=0,padx=5,pady=5)
        address.grid(row=4,column=0,padx=5,pady=5)
        number.grid(row=5,column=0,padx=5,pady=5)
        date.grid(row=6,column=0,padx=5,pady=5)
        time.grid(row=7,column=0,padx=5,pady=5)

# entryID=Entry(root,width=25,bd=5,font=("Times New Roman",15))
        entryName=Entry(root5,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryAge=Entry(root5,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryAddress=Entry(root5,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryNumber=Entry(root5,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryDate=Entry(root5,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryTime=Entry(root5,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)

# entryID.grid(row=1,column=1,columnspan=3,padx=5,pady=5)
        entryName.grid(row=2,column=1,columnspan=3,padx=5,pady=5)
        entryAge.grid(row=3,column=1,columnspan=3,padx=5,pady=5)
        entryAddress.grid(row=4,column=1,columnspan=3,padx=5,pady=5)
        entryNumber.grid(row=5,column=1,columnspan=3,padx=5,pady=5)
        entryDate.grid(row=6,column=1,columnspan=3,padx=5,pady=5)
        entryTime.grid(row=7,column=1,columnspan=3,padx=5,pady=5)

       

        submit=Button(root5,text="Submit",padx=1,pady=1,width=5,bd=1,font=("Times New Roman",15),bg="RED",fg="white",command=insert_data)
        submit.grid(row=8,column=1)

        update=Button(root5,text="Update",padx=1,pady=1,width=5,bd=1,font=("Times New Roman",15),bg="Red",fg="white",command=update_data)
        update.grid(row=8,column=2)

        delete=Button(root5,text="Delete",padx=1,pady=1,width=5,bd=1,font=("Times New Roman",15),bg="Red",fg="white",command=delete_data)
        delete.grid(row=8,column=3)

#####for treeview
        style=ttk.Style()
        style.configure("Treeview.Heading",font=("Times New Roman",15))


        win["columns"]=("ID","Name","Age","Address","Number","Date","Time")
        win.column("#0", width=0,stretch=NO)
        win.column("ID",anchor=W,width=30)
        win.column("Name",anchor=W,width=100)
        win.column("Age",anchor=W,width=50)
        win.column("Address",anchor=W,width=90)
        win.column("Number",anchor=W,width=90)
        win.column("Date",anchor=W,width=60)
        win.column("Time",anchor=W,width=60)

# win.heading("#0", width=0,stretch=NO)
        win.heading("ID",text="ID",anchor=W)
        win.heading("Name",text="Name",anchor=W)
        win.heading("Age",text="Age",anchor=W)
        win.heading("Address",text="Address",anchor=W)
        win.heading("Number",text="Number",anchor=W)
        win.heading("Date",text="Date",anchor=W)
        win.heading("Time",text="Time",anchor=W)

        for data in win.get_children():
            win.delete(data)
        for result in reverse(read()):
            win.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
    
    
        win.tag_configure("orow",background="Red",font=("Times New Roman",8))
        win.grid(row=1,column=4,columnspan=4,rowspan=10,padx=10,pady=10)

        root5.mainloop()

#######################Norvic Hospital
    def listed():
        root3.destroy()
        root6=Tk()
        root6.title("Norvic Appointment Booking System",)
        root6.geometry("1000x500")
        root6.iconbitmap("88.ico")
        root6.configure(bg="Gray")

        WON=ttk.Treeview(root6)
        Hospital="Norvic Hospital"

        def reverse(tuples):
            new=tuples[::-1]
            return new

        
        def insert(id,name,age,address,number,date,time):
                database.addAppointment(name,age,address,number,date,time,4,user.username)
                messagebox.showinfo("Norvichospital","Booked sucessfully")
    

        def delete_data_database(data):
            database.deleteAppointment(int(data))
            messagebox.showinfo("Norvichospital","Deleted sucessfully")

        def update_database(id,name,age,address,number,date,time):
            database.updateAppointment(int(id),name,age,address,number,date,time)
            messagebox.showinfo("Norvichospital","Updated sucessfully")

        def read():
                if(not user.isAdmin):
                    return database.getAppointmentListForUser(4,user.username)
                else:
                   return database.getAppointmentListForAdmin(4)

        def insert_data():

#   id=str(entryID.get())
            name=str(entryName.get())
            age=str(entryAge.get())
            address=str(entryAddress.get())
            number=str(entryNumber.get())
            date=str(entryDate.get())
            time=str(entryTime.get())

            if id==""or id==" ":
                print("Error Inserting Id")
            if name==""or name==" ":
                print("Error Inserting Name") 
            if age==""or age==" ":
                print("Error Inserting Age") 
            if address==""or address==" ":
                print("Error Inserting Address")
            if number==""or number==" ":
                print("Error Inserting Number") 
            if date==""or date==" ":
                print("Error Inserting Date") 
            if time==""or time==" ":
                print("Error Inserting Time")

            else:
                insert(str(id),str(name),str(age),str(address),str(number),str(date),str(time))

            for data in WON.get_children():
                 WON.delete(data)
            for result in reverse(read()):
                 WON.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))


    
            WON.tag_configure("orow",background="Red",font=("Times New Roman",8))
            WON.grid(row=1,column=4,columnspan=4,rowspan=5,padx=10,pady=10)
 

        def delete_data():
                 selected_app=WON.selection()[0]
                 deleteData=str(WON.item(selected_app)["values"][0])
                 delete_data_database(deleteData)

                 for data in WON.get_children():
                  WON.delete(data)
                 for result in reverse(read()):
                  WON.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
    
                 WON.tag_configure("orow",background="Red",font=("Times New Roman",8))
                 WON.grid(row=1,column=4,columnspan=4,rowspan=5,padx=10,pady=10) 
      
        def update_data():             
                if(len(WON.selection())<1): return
                selected_app=WON.selection()[0]
                updatedata=str(WON.item(selected_app)["values"][0])
                update_database(updatedata,entryName.get(),entryAge.get(),entryAddress.get(),entryNumber.get(),entryDate.get(),entryTime.get())

                for data in WON.get_children():
                  WON.delete(data)
                for result in reverse(read()):
                  WON.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
    
                WON.tag_configure("orow",background="Red",font=("Times New Roman",8))
                WON.grid(row=1,column=4,columnspan=4,rowspan=5,padx=10,pady=10)

    


        title=Label(root6,text="Norvic Appointment",font=("Times New Roman",30),bd=2)
        title.grid(row=0,column=0,columnspan=8,padx=20,pady=20)

# id=Label(root,text="ID",font=("Times New Roman",15))
        name=Label(root6,text="Name",font=("Times New Roman",17),bg="gray",fg="white")
        age=Label(root6,text="Age",font=("Times New Roman",17),bg="gray",fg="white")
        address=Label(root6,text="Address",font=("Times New Roman",17),bg="gray",fg="white")
        number=Label(root6,text="Phone Number",font=("Times New Roman",17),bg="gray",fg="white")
        date=Label(root6,text="Date",font=("Times New Roman",17),bg="gray",fg="white")
        time=Label(root6,text="Time",font=("Times New Roman",17),bg="gray",fg="white")

# id.grid(row=1,column=0,padx=5,pady=5)
        name.grid(row=2,column=0,padx=5,pady=5)
        age.grid(row=3,column=0,padx=5,pady=5)
        address.grid(row=4,column=0,padx=5,pady=5)
        number.grid(row=5,column=0,padx=5,pady=5)
        date.grid(row=6,column=0,padx=5,pady=5)
        time.grid(row=7,column=0,padx=5,pady=5)

# entryID=Entry(root,width=25,bd=5,font=("Times New Roman",15))
        entryName=Entry(root6,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryAge=Entry(root6,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryAddress=Entry(root6,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryNumber=Entry(root6,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryDate=Entry(root6,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryTime=Entry(root6,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)

# entryID.grid(row=1,column=1,columnspan=3,padx=5,pady=5)
        entryName.grid(row=2,column=1,columnspan=3,padx=5,pady=5)
        entryAge.grid(row=3,column=1,columnspan=3,padx=5,pady=5)
        entryAddress.grid(row=4,column=1,columnspan=3,padx=5,pady=5)
        entryNumber.grid(row=5,column=1,columnspan=3,padx=5,pady=5)
        entryDate.grid(row=6,column=1,columnspan=3,padx=5,pady=5)
        entryTime.grid(row=7,column=1,columnspan=3,padx=5,pady=5)

        submit=Button(root6,text="Submit",padx=1,pady=1,width=5,bd=1,font=("Times New Roman",15),bg="RED",fg="white",command=insert_data)
        submit.grid(row=8,column=1)

        update=Button(root6,text="Update",padx=1,pady=1,width=5,bd=1,font=("Times New Roman",15),bg="Red",fg="white",command=update_data)
        update.grid(row=8,column=2)

        delete=Button(root6,text="Delete",padx=1,pady=1,width=5,bd=1,font=("Times New Roman",15),bg="Red",fg="white",command=delete_data)
        delete.grid(row=8,column=3)

#####for treeview
        style=ttk.Style()
        style.configure("Treeview.Heading",font=("Times New Roman",15))


        WON["columns"]=("ID","Name","Age","Address","Number","Date","Time")
        WON.column("#0", width=0,stretch=NO)
        WON.column("ID",anchor=W,width=30)
        WON.column("Name",anchor=W,width=100)
        WON.column("Age",anchor=W,width=50)
        WON.column("Address",anchor=W,width=90)
        WON.column("Number",anchor=W,width=90)
        WON.column("Date",anchor=W,width=60)
        WON.column("Time",anchor=W,width=60)

# WON.heading("#0", width=0,stretch=NO)
        WON.heading("ID",text="ID",anchor=W)
        WON.heading("Name",text="Name",anchor=W)
        WON.heading("Age",text="Age",anchor=W)
        WON.heading("Address",text="Address",anchor=W)
        WON.heading("Number",text="Number",anchor=W)
        WON.heading("Date",text="Date",anchor=W)
        WON.heading("Time",text="Time",anchor=W)

        for data in WON.get_children():
            WON.delete(data)
        for result in reverse(read()):
            WON.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
    
    
        WON.tag_configure("orow",background="Red",font=("Times New Roman",8))
        WON.grid(row=1,column=4,columnspan=4,rowspan=10,padx=10,pady=10) 


        root6.mainloop()

###############BIr hospital################
    def hooked():
        root3.destroy()
        root7=Tk()
        root7.title("Bir Appointment Booking System",)
        root7.geometry("1000x500")
        root7.iconbitmap("88.ico")
        root7.configure(bg="Gray")

        new=ttk.Treeview(root7)
        Hospital="Bir Hospital"

        def reverse(tuples):
            new=tuples[::-1]
            return new

        def insert(id,name,age,address,number,date,time):
                database.addAppointment(name,age,address,number,date,time,5,user.username)
                messagebox.showinfo("Bir hospital","Booked sucessfully")
    

        def delete_data_database(data):
                database.deleteAppointment(int(data))
                messagebox.showinfo("Bir hospital","Deleted sucessfully")

        def update_database(id,name,age,address,number,date,time):
                database.updateAppointment(int(id),name,age,address,number,date,time)
                messagebox.showinfo("Birhospital","Updated sucessfully")

        def read():
                if(not user.isAdmin):
                    return database.getAppointmentListForUser(5,user.username)
                else:
                   return database.getAppointmentListForAdmin(5)

        def insert_data():

#   id=str(entryID.get())
                 name=str(entryName.get())
                 age=str(entryAge.get())
                 address=str(entryAddress.get())
                 number=str(entryNumber.get())
                 date=str(entryDate.get())
                 time=str(entryTime.get())

                 if id==""or id==" ":
                     print("Error Inserting Id")
                 if name==""or name==" ":
                     print("Error Inserting Name") 
                 if age==""or age==" ":
                       print("Error Inserting Age") 
                 if address==""or address==" ":
                     print("Error Inserting Address")
                 if number==""or number==" ":
                     print("Error Inserting Number") 
                 if date==""or date==" ":
                     print("Error Inserting Date") 
                 if time==""or time==" ":
                     print("Error Inserting Time")

                 else:
                  insert(str(id),str(name),str(age),str(address),str(number),str(date),str(time))

                 for data in new.get_children():
                    new.delete(data)
                 for result in reverse(read()):
                    new.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))


    
                    new.tag_configure("orow",background="Red",font=("Times New Roman",8))
                    new.grid(row=1,column=4,columnspan=4,rowspan=5,padx=10,pady=10)
 

        def delete_data():
            selected_app=new.selection()[0]
            deleteData=str(new.item(selected_app)["values"][0])
            delete_data_database(deleteData)

            for data in new.get_children():
                 new.delete(data)
            for result in reverse(read()):
                  new.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
    
            new.tag_configure("orow",background="Red",font=("Times New Roman",8))
            new.grid(row=1,column=4,columnspan=4,rowspan=5,padx=10,pady=10) 
      
        def update_data():             
              if(len(new.selection())<1): return
              selected_app=new.selection()[0]
              updatedata=str(new.item(selected_app)["values"][0])
              update_database(updatedata,entryName.get(),entryAge.get(),entryAddress.get(),entryNumber.get(),entryDate.get(),entryTime.get())

              for data in new.get_children():
                  new.delete(data)
              for result in reverse(read()):
                  new.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
    
                  new.tag_configure("orow",background="Red",font=("Times New Roman",8))
                  new.grid(row=1,column=4,columnspan=4,rowspan=5,padx=10,pady=10)

    


        title=Label(root7,text="Bir Hospital Appointment",font=("Times New Roman",30),bd=2)
        title.grid(row=0,column=0,columnspan=8,padx=20,pady=20)

# id=Label(root,text="ID",font=("Times New Roman",15))
        name=Label(root7,text="Name",font=("Times New Roman",17),bg="gray",fg="white")
        age=Label(root7,text="Age",font=("Times New Roman",17),bg="gray",fg="white")
        address=Label(root7,text="Address",font=("Times New Roman",17),bg="gray",fg="white")
        number=Label(root7,text="Phone Number",font=("Times New Roman",17),bg="gray",fg="white")
        date=Label(root7,text="Date",font=("Times New Roman",17),bg="gray",fg="white")
        time=Label(root7,text="Time",font=("Times New Roman",17),bg="gray",fg="white")

# id.grid(row=1,column=0,padx=5,pady=5)
        name.grid(row=2,column=0,padx=5,pady=5)
        age.grid(row=3,column=0,padx=5,pady=5)
        address.grid(row=4,column=0,padx=5,pady=5)
        number.grid(row=5,column=0,padx=5,pady=5)
        date.grid(row=6,column=0,padx=5,pady=5)
        time.grid(row=7,column=0,padx=5,pady=5)

# entryID=Entry(root,width=25,bd=5,font=("Times New Roman",15))
        entryName=Entry(root7,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryAge=Entry(root7,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryAddress=Entry(root7,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryNumber=Entry(root7,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryDate=Entry(root7,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryTime=Entry(root7,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)

# entryID.grid(row=1,column=1,columnspan=3,padx=5,pady=5)
        entryName.grid(row=2,column=1,columnspan=3,padx=5,pady=5)
        entryAge.grid(row=3,column=1,columnspan=3,padx=5,pady=5)
        entryAddress.grid(row=4,column=1,columnspan=3,padx=5,pady=5)
        entryNumber.grid(row=5,column=1,columnspan=3,padx=5,pady=5)
        entryDate.grid(row=6,column=1,columnspan=3,padx=5,pady=5)
        entryTime.grid(row=7,column=1,columnspan=3,padx=5,pady=5)

        submit=Button(root7,text="Submit",padx=1,pady=1,width=5,bd=1,font=("Times New Roman",15),bg="RED",fg="white",command=insert_data)
        submit.grid(row=8,column=1)

        update=Button(root7,text="Update",padx=1,pady=1,width=5,bd=1,font=("Times New Roman",15),bg="Red",fg="white",command=update_data)
        update.grid(row=8,column=2)

        delete=Button(root7,text="Delete",padx=1,pady=1,width=5,bd=1,font=("Times New Roman",15),bg="Red",fg="white",command=delete_data)
        delete.grid(row=8,column=3)

#####for treeview
        style=ttk.Style()
        style.configure("Treeview.Heading",font=("Times New Roman",15))


        new["columns"]=("ID","Name","Age","Address","Number","Date","Time")
        new.column("#0", width=0,stretch=NO)
        new.column("ID",anchor=W,width=30)
        new.column("Name",anchor=W,width=100)
        new.column("Age",anchor=W,width=50)
        new.column("Address",anchor=W,width=90)
        new.column("Number",anchor=W,width=90)
        new.column("Date",anchor=W,width=60)
        new.column("Time",anchor=W,width=60)

# new.heading("#0", width=0,stretch=NO)
        new.heading("ID",text="ID",anchor=W)
        new.heading("Name",text="Name",anchor=W)
        new.heading("Age",text="Age",anchor=W)
        new.heading("Address",text="Address",anchor=W)
        new.heading("Number",text="Number",anchor=W)
        new.heading("Date",text="Date",anchor=W)
        new.heading("Time",text="Time",anchor=W)

        for data in new.get_children():
            new.delete(data)
        for result in reverse(read()):
              new.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
    
    
        new.tag_configure("orow",background="Red",font=("Times New Roman",8))
        new.grid(row=1,column=4,columnspan=4,rowspan=10,padx=10,pady=10) 








        root7.mainloop()

####################################Grrandy Hospital#########
    def saved():
            root3.destroy()
            root8=Tk()
            root8.title("Grandy Appointment Booking System",)
            root8.geometry("1000x500")
            root8.iconbitmap("88.ico")
            root8.configure(bg="Gray")

            rip=ttk.Treeview(root8)
            Hospital="Grandy Hospital"

            def reverse(tuples):
                rip=tuples[::-1]
                return rip

            def insert(id,name,age,address,number,date,time):
                database.addAppointment(name,age,address,number,date,time,2,user.username)
                messagebox.showinfo("Grandy hospital","Booked sucessfully")
    

            def delete_data_database(data):
                database.deleteAppointment(int(data))
                messagebox.showinfo("Grandy hospital","Deleted sucessfully")

            def update_database(id,name,age,address,number,date,time):
               database.updateAppointment(int(id),name,age,address,number,date,time)
               messagebox.showinfo("Grandy hospital","Updated sucessfully")

            def read():
                 if(not user.isAdmin):
                    return database.getAppointmentListForUser(2,user.username)
                 else:
                   return database.getAppointmentListForAdmin(2)

            def insert_data():

#   id=str(entryID.get())
                name=str(entryName.get())
                age=str(entryAge.get())
                address=str(entryAddress.get())
                number=str(entryNumber.get())
                date=str(entryDate.get())
                time=str(entryTime.get())

                if id==""or id==" ":
                    print("Error Inserting Id")
                if name==""or name==" ":
                    print("Error Inserting Name") 
                if age==""or age==" ":
                    print("Error Inserting Age") 
                if address==""or address==" ":
                    print("Error Inserting Address")
                if number==""or number==" ":
                    print("Error Inserting Number") 
                if date==""or date==" ":
                   print("Error Inserting Date") 
                if time==""or time==" ":
                   print("Error Inserting Time")

                else:
                    insert(str(id),str(name),str(age),str(address),str(number),str(date),str(time))

                for data in rip.get_children():
                     rip.delete(data)
                for result in reverse(read()):
                     rip.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))


    
                     rip.tag_configure("orow",background="Red",font=("Times rip Roman",8))
                     rip.grid(row=1,column=4,columnspan=4,rowspan=5,padx=10,pady=10)
 

            def delete_data():
             selected_app=rip.selection()[0]
             deleteData=str(rip.item(selected_app)["values"][0])
             delete_data_database(deleteData)

             for data in rip.get_children():
                 rip.delete(data)
             for result in reverse(read()):
                 rip.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
    
             rip.tag_configure("orow",background="Red",font=("Times rip Roman",8))
             rip.grid(row=1,column=4,columnspan=4,rowspan=5,padx=10,pady=10) 
      
            def update_data():             
                if(len(rip.selection())<1): return
                selected_app=rip.selection()[0]
                updatedata=str(rip.item(selected_app)["values"][0])
                update_database(updatedata,entryName.get(),entryAge.get(),entryAddress.get(),entryNumber.get(),entryDate.get(),entryTime.get())

                for data in rip.get_children():
                  rip.delete(data)
                for result in reverse(read()):
                  rip.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
    
            rip.tag_configure("orow",background="Red",font=("Times rip Roman",8))
            rip.grid(row=1,column=4,columnspan=4,rowspan=5,padx=10,pady=10)

    


            title=Label(root8,text="Grandy Hospital Appointment",font=("Times rip Roman",30),bd=2)
            title.grid(row=0,column=0,columnspan=8,padx=20,pady=20)

# id=Label(root,text="ID",font=("Times rip Roman",15))
            name=Label(root8,text="Name",font=("Times rip Roman",17),bg="gray",fg="white")
            age=Label(root8,text="Age",font=("Times rip Roman",17),bg="gray",fg="white")
            address=Label(root8,text="Address",font=("Times rip Roman",17),bg="gray",fg="white")
            number=Label(root8,text="Phone Number",font=("Times rip Roman",17),bg="gray",fg="white")
            date=Label(root8,text="Date",font=("Times rip Roman",17),bg="gray",fg="white")
            time=Label(root8,text="Time",font=("Times rip Roman",17),bg="gray",fg="white")

# id.grid(row=1,column=0,padx=5,pady=5)
            name.grid(row=2,column=0,padx=5,pady=5)
            age.grid(row=3,column=0,padx=5,pady=5)
            address.grid(row=4,column=0,padx=5,pady=5)
            number.grid(row=5,column=0,padx=5,pady=5)
            date.grid(row=6,column=0,padx=5,pady=5)
            time.grid(row=7,column=0,padx=5,pady=5)

# entryID=Entry(root,width=25,bd=5,font=("Times rip Roman",15))
            entryName=Entry(root8,width=25,bd=2,font=("Times rip Roman",15),highlightbackground="white",highlightthickness=3)
            entryAge=Entry(root8,width=25,bd=2,font=("Times rip Roman",15),highlightbackground="white",highlightthickness=3)
            entryAddress=Entry(root8,width=25,bd=2,font=("Times rip Roman",15),highlightbackground="white",highlightthickness=3)
            entryNumber=Entry(root8,width=25,bd=2,font=("Times rip Roman",15),highlightbackground="white",highlightthickness=3)
            entryDate=Entry(root8,width=25,bd=2,font=("Times rip Roman",15),highlightbackground="white",highlightthickness=3)
            entryTime=Entry(root8,width=25,bd=2,font=("Times rip Roman",15),highlightbackground="white",highlightthickness=3)

# entryID.grid(row=1,column=1,columnspan=3,padx=5,pady=5)
            entryName.grid(row=2,column=1,columnspan=3,padx=5,pady=5)
            entryAge.grid(row=3,column=1,columnspan=3,padx=5,pady=5)
            entryAddress.grid(row=4,column=1,columnspan=3,padx=5,pady=5)
            entryNumber.grid(row=5,column=1,columnspan=3,padx=5,pady=5)
            entryDate.grid(row=6,column=1,columnspan=3,padx=5,pady=5)
            entryTime.grid(row=7,column=1,columnspan=3,padx=5,pady=5)

            submit=Button(root8,text="Submit",padx=1,pady=1,width=5,bd=1,font=("Times rip Roman",15),bg="RED",fg="white",command=insert_data)
            submit.grid(row=8,column=1)

            update=Button(root8,text="Update",padx=1,pady=1,width=5,bd=1,font=("Times rip Roman",15),bg="Red",fg="white",command=update_data)
            update.grid(row=8,column=2)

            delete=Button(root8,text="Delete",padx=1,pady=1,width=5,bd=1,font=("Times rip Roman",15),bg="Red",fg="white",command=delete_data)
            delete.grid(row=8,column=3)

#####for treeview
            style=ttk.Style()
            style.configure("Treeview.Heading",font=("Times rip Roman",15))


            rip["columns"]=("ID","Name","Age","Address","Number","Date","Time")
            rip.column("#0", width=0,stretch=NO)
            rip.column("ID",anchor=W,width=30)
            rip.column("Name",anchor=W,width=100)
            rip.column("Age",anchor=W,width=50)
            rip.column("Address",anchor=W,width=90)
            rip.column("Number",anchor=W,width=90)
            rip.column("Date",anchor=W,width=60)
            rip.column("Time",anchor=W,width=60)

# rip.heading("#0", width=0,stretch=NO)
            rip.heading("ID",text="ID",anchor=W)
            rip.heading("Name",text="Name",anchor=W)
            rip.heading("Age",text="Age",anchor=W)
            rip.heading("Address",text="Address",anchor=W)
            rip.heading("Number",text="Number",anchor=W)
            rip.heading("Date",text="Date",anchor=W)
            rip.heading("Time",text="Time",anchor=W)

            for data in rip.get_children():
                rip.delete(data)
            for result in reverse(read()):
                rip.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
    
    
            rip.tag_configure("orow",background="Red",font=("Times rip Roman",8))
            rip.grid(row=1,column=4,columnspan=4,rowspan=10,padx=10,pady=10) 


            root8.mainloop()

########################################################


    frame11=Frame(root3,width=80,height=80,bg="#fff")
    frame11.place(x=15,y=90)
    frame2=Frame(root3,width=75,height=20,bg="white")
    frame2.place(x=15,y=185)
    button_1=Button(frame2,text="Om Hospital",fg="white",bg="red",font=("Times New Roman",10,"bold"),command=appointment)
    button_1.place(x=0,y=0)
    global image_var  # Declare as a global variable
    image_var = ImageTk.PhotoImage(Image.open("om.png"))
    abclbl = Label(frame11, image=image_var)
    abclbl.image = image_var  # Store the reference
    abclbl.pack()
    frame12=Frame(root3,width=80,height=80,bg="#fff")
    frame12.place(x=250,y=90)
    frame3=Frame(root3,width=118,height=20,bg="white")
    frame3.place(x=235,y=175)
    button_2=Button(frame3,text="Manmohan Hospital",fg="white",bg="red",font=("Times New Roman",10,"bold"),command=booking)
    button_2.place(x=0,y=0)
    # global image_ch  # Declare as a global variable
    image_ch = ImageTk.PhotoImage(Image.open("manmohan.png"))
    abclbl = Label(frame12, image=image_ch)
    abclbl.image = image_ch  # Store the reference
    abclbl.pack()
    frame13=Frame(root3,width=80,height=80,bg="#fff")
    frame13.place(x=490,y=90)
    frame4=Frame(root3,width=95,height=20,bg="white")
    frame4.place(x=485,y=175)
    button_2=Button(frame4,text="Norvic Hospital",fg="white",bg="red",font=("Times New Roman",10,"bold"),command=listed)
    button_2.place(x=0,y=0)
    # global image_ch  # Declare as a global variable
    image_abh = ImageTk.PhotoImage(Image.open("Norvic.png"))
    abclbl = Label(frame13, image=image_abh)
    abclbl.image = image_abh  # Store the reference
    abclbl.pack()
    frame14=Frame(root3,width=80,height=80,bg="#fff")
    frame14.place(x=730,y=90)
    frame5=Frame(root3,width=78,height=20,bg="white")
    frame5.place(x=735,y=190)
    button_2=Button(frame5,text="Bir Hospital",fg="white",bg="red",font=("Times New Roman",10,"bold"),command=hooked)
    button_2.place(x=0,y=0)
    # global image_ch  # Declare as a global variable
    image_abc = ImageTk.PhotoImage(Image.open("bir.png"))
    abclbl = Label(frame14, image=image_abc)
    abclbl.image = image_abc  # Store the reference
    abclbl.pack()
    frame15=Frame(root3,width=80,height=80,bg="#fff")
    frame15.place(x=18,y=290)
    frame6=Frame(root3,width=95,height=20,bg="white")
    frame6.place(x=15,y=375)
    button_2=Button(frame6,text="Grandy Hospital",fg="white",bg="red",font=("Times New Roman",10,"bold"),command=saved)
    button_2.place(x=0,y=0)
    # global image_ch  # Declare as a global variable
    image_abc = ImageTk.PhotoImage(Image.open("grandy.png"))
    abclbl = Label(frame15, image=image_abc)
    abclbl.image = image_abc  # Store the reference
    abclbl.pack()
    root3.mainloop()

##################First aid#####################

def first_aid():
    aid = Tk()
    aid.title("First Aid")
    aid.geometry("650x660")
    aid.iconbitmap("v.ico")
    aid.resizable(0, 0)
    aid.configure(bg="white")


    def click():
        aid.destroy()
        first = Tk()
        first.title("First Aid Booking System",)
        first.geometry("1000x500")
        first.configure(bg="#FF0000")

        win = ttk.Treeview(first)
        Hospital = "First Aid"

        def reverse(tuples):
            new = tuples[::-1]
            return new

        def insert( name, address, number, quantity):
            firstaid.addbooking(name,address,number,quantity,user.username)
            messagebox.showinfo("First Aid","Booked sucessfully")

        def delete_data_database(data):
            firstaid.deletebooking(int(data))
            messagebox.showinfo("First Aid","Deleted sucessfully")
            

        def update_database(id, name, address, number, quantity):
            firstaid.updatebooking(int(id),name,address,number,quantity)
            messagebox.showinfo("First Aid","Updated sucessfully")
            

        def read():
            if(not user.isAdmin):
                return firstaid.getbookingListForUser(user.username)
            else:
                return firstaid.getbookingListForAdmin()

        def insert_data():

            #   id=str(entryID.get())
            name = str(entryName.get())
            address = str(entryAddress.get())
            number = str(entryNumber.get())
            quantity = str(entryQuantity.get())

            if id == "" or id == " ":
                print("Error Inserting Id")
            if name == "" or name == " ":
                print("Error Inserting Name")
            if address == "" or address == " ":
                print("Error Inserting Address")
            if number == "" or number == " ":
                print("Error Inserting Number")
            if quantity == "" or quantity == " ":
                print("Error Inserting Quantity")

            else:
                insert(str(name), str(address), str(number), str(quantity))

            for data in win.get_children():
                win.delete(data)
            for result in reverse(read()):
                win.insert(parent="", index="end", iid=result[0], text="", values=(
                    result), tag=str(result[0]))

            win.tag_configure("orow", background="Red",
                            font=("Times New Roman", 8))
            win.grid(row=1, column=4, columnspan=4, rowspan=5, padx=10, pady=10)

        def delete_data():
            selected_app = win.selection()[0]
            deleteData = str(win.item(selected_app)["values"][0])
            delete_data_database(deleteData)

            for data in win.get_children():
                win.delete(data)

            for result in reverse(read()):
                win.insert(parent="", index="end", iid=result[0], text="", values=(
                    result), tag=str(result[0]))

            win.tag_configure("orow", background="Red",
                            font=("Times New Roman", 8))
            win.grid(row=1, column=4, columnspan=4, rowspan=5, padx=10, pady=10)

        def update_data():
            if (len(win.selection()) < 1):
                return
            selected_app = win.selection()[0]
            updatedata = str(win.item(selected_app)["values"][0])
            update_database(updatedata, entryName.get(), entryAddress.get(),
                            entryNumber.get(), entryQuantity.get())

            for data in win.get_children():
                win.delete(data)
            for result in reverse(read()):
                win.insert(parent="", index="end", iid=result[0], text="", values=(
                    result), tag=str(result[0]))

            win.tag_configure("orow", background="Red",
                            font=("Times New Roman", 8))
            win.grid(row=1, column=4, columnspan=4, rowspan=5, padx=10, pady=10)

        title = Label(first, text="First Aid",
                    font=("Times New Roman", 30), bd=2)
        title.grid(row=0, column=0, columnspan=8, padx=20, pady=20)

    # id=Label(aid2,text="ID",font=("Times New Roman",15))
        name = Label(first, text="Name", font=(
            "Times New Roman", 17), bg="#FF0000", fg="white")
        address = Label(first, text="Address", font=(
            "Times New Roman", 17), bg="#FF0000", fg="white")
        number = Label(first, text="Phone Number", font=(
            "Times New Roman", 17), bg="#FF0000", fg="white")
        quantity = Label(first, text="Quantity", font=(
            "Times New Roman", 17), bg="#FF0000", fg="white")


    # id.grid(row=1,column=0,padx=5,pady=5)
        name.grid(row=2, column=0, padx=5, pady=5)
        address.grid(row=3, column=0, padx=5, pady=5)
        number.grid(row=4, column=0, padx=5, pady=5)
        quantity.grid(row=5, column=0, padx=5, pady=5)


    # entryID=Entry(aid2,width=25,bd=5,font=("Times New Roman",15))
        entryName = Entry(first, width=25, bd=2, font=(
            "Times New Roman", 15), highlightbackground="white", highlightthickness=3)
        entryAddress = Entry(first, width=25, bd=2, font=(
            "Times New Roman", 15), highlightbackground="white", highlightthickness=3)
        entryNumber = Entry(first, width=25, bd=2, font=(
            "Times New Roman", 15), highlightbackground="white", highlightthickness=3)
        entryQuantity = Entry(first, width=25, bd=2, font=(
            "Times New Roman", 15), highlightbackground="white", highlightthickness=3)


    # entryID.grid(row=1,column=1,columnspan=3,padx=5,pady=5)
        entryName.grid(row=2, column=1, columnspan=3, padx=5, pady=5)
        entryAddress.grid(row=3, column=1, columnspan=3, padx=5, pady=5)
        entryNumber.grid(row=4, column=1, columnspan=3, padx=5, pady=5)
        entryQuantity.grid(row=5, column=1, columnspan=3, padx=5, pady=5)

        submit = Button(first, text="Submit", padx=1, pady=1, width=5, bd=1, font=(
            "Times New Roman", 15), bg="RED", fg="white", command=insert_data)
        submit.grid(row=8, column=1)

        update = Button(first, text="Update", padx=1, pady=1, width=5, bd=1, font=(
            "Times New Roman", 15), bg="Red", fg="white", command=update_data)
        update.grid(row=8, column=2)

        delete = Button(first, text="Delete", padx=1, pady=1, width=5, bd=1, font=(
            "Times New Roman", 15), bg="Red", fg="white", command=delete_data)
        delete.grid(row=8, column=3)

    # for treeview
        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Times New Roman", 15))

        win["columns"] = ("ID", "Name", "Address", "Number", "Quantity")
        win.column("#0", width=0, stretch=NO)
        win.column("ID", anchor=W, width=30)
        win.column("Name", anchor=W, width=100)
        win.column("Address", anchor=W, width=100)
        win.column("Number", anchor=W, width=90)
        win.column("Quantity", anchor=W, width=90)


    # win.heading("#0", width=0,stretch=NO)
        win.heading("ID", text="ID", anchor=W)
        win.heading("Name", text="Name", anchor=W)
        win.heading("Address", text="Address", anchor=W)
        win.heading("Number", text="Number", anchor=W)
        win.heading("Quantity", text="Quantity", anchor=W)

        for data in win.get_children():
            win.delete(data)
        for result in reverse(read()):
            win.insert(parent="", index="end", iid=result[0], text="", values=(result), tag=str(result[0]))

        win.tag_configure("orow", background="Red", font=("Times New Roman", 8))
        win.grid(row=1, column=4, columnspan=4, rowspan=10, padx=10, pady=10)

        first.mainloop()


    Frame
    frame0 = Frame(aid, height="40", width="50", bg="white",)
    frame0.place(x=5, y=120)
    frame1 = Frame(aid, height="100", width="200", bg="red",)
    frame1.pack(fill="both")  # .pack(fill="both")means it fills colour
    frame2 = Frame(aid, height="40", width="150", bg="white",)
    frame2.place(x=80, y=120)
    frame3 = Frame(aid, height="40", width="180", bg="white",)
    frame3.place(x=280, y=120)
    frame4 = Frame(aid, height="40", width="100", bg="white",)
    frame4.place(x=500, y=120)
    frame5 = Frame(aid, height="40", width="50", bg="white",)
    frame5.place(x=5, y=170)
    frame6 = Frame(aid, height="40", width="150", bg="white",)
    frame6.place(x=80, y=170)
    frame7 = Frame(aid, height="40", width="180", bg="white",)
    frame7.place(x=280, y=170)
    frame8 = Frame(aid, height="40", width="100", bg="white",)
    frame8.place(x=500, y=170)
    frame9 = Frame(aid, height="40", width="50", bg="white",)
    frame9.place(x=5, y=220)
    frame10 = Frame(aid, height="40", width="150", bg="white",)
    frame10.place(x=80, y=220)
    frame11 = Frame(aid, height="40", width="180", bg="white",)
    frame11.place(x=280, y=220)
    frame12 = Frame(aid, height="40", width="100", bg="white",)
    frame12.place(x=500, y=220)
    frame13 = Frame(aid, height="40", width="50", bg="white",)
    frame13.place(x=5, y=270)
    frame14 = Frame(aid, height="40", width="150", bg="white",)
    frame14.place(x=80, y=270)
    frame15 = Frame(aid, height="40", width="180", bg="white",)
    frame15.place(x=280, y=270)
    frame16 = Frame(aid, height="40", width="100", bg="white",)
    frame16.place(x=500, y=270)
    frame17 = Frame(aid, height="40", width="50", bg="white",)
    frame17.place(x=5, y=320)
    frame18 = Frame(aid, height="40", width="150", bg="white",)
    frame18.place(x=80, y=320)
    frame19 = Frame(aid, height="40", width="180", bg="white",)
    frame19.place(x=280, y=320)
    frame20 = Frame(aid, height="40", width="100", bg="white",)
    frame20.place(x=500, y=320)
    frame21 = Frame(aid, height="40", width="50", bg="white",)
    frame21.place(x=5, y=370)
    frame22 = Frame(aid, height="40", width="150", bg="white",)
    frame22.place(x=80, y=370)
    frame23 = Frame(aid, height="40", width="180", bg="white",)
    frame23.place(x=280, y=370)
    frame24 = Frame(aid, height="40", width="100", bg="white",)
    frame24.place(x=500, y=370)
    frame25 = Frame(aid, height="40", width="50", bg="white",)
    frame25.place(x=5, y=420)
    frame26 = Frame(aid, height="40", width="150", bg="white",)
    frame26.place(x=80, y=420)
    frame27 = Frame(aid, height="40", width="180", bg="white",)
    frame27.place(x=280, y=420)
    frame28 = Frame(aid, height="40", width="100", bg="white",)
    frame28.place(x=500, y=420)
    frame29 = Frame(aid, height="40", width="50", bg="white",)
    frame29.place(x=5, y=470)
    frame30 = Frame(aid, height="40", width="190", bg="white",)
    frame30.place(x=80, y=470)
    frame31 = Frame(aid, height="40", width="180", bg="white",)
    frame31.place(x=280, y=470)
    frame32 = Frame(aid, height="40", width="100", bg="white",)
    frame32.place(x=500, y=470)
    frame33 = Frame(aid, height="40", width="50", bg="white",)
    frame33.place(x=5, y=520)
    frame34 = Frame(aid, height="40", width="150", bg="white",)
    frame34.place(x=80, y=520)
    frame35 = Frame(aid, height="40", width="180", bg="white",)
    frame35.place(x=280, y=520)
    frame36 = Frame(aid, height="40", width="100", bg="white",)
    frame36.place(x=500, y=520)
    frame37 = Frame(aid, height="40", width="50", bg="white",)
    frame37.place(x=5, y=570)
    frame38 = Frame(aid, height="40", width="168", bg="white",)
    frame38.place(x=80, y=570)
    frame39 = Frame(aid, height="40", width="180", bg="white",)
    frame39.place(x=280, y=570)
    frame40 = Frame(aid, height="40", width="100", bg="white",)
    frame40.place(x=500, y=570)
    frame41 = Frame(aid, height="40", width="650", bg="red",)
    frame41.place(x=0, y=620)


    # Label


    label = Label(frame0, text="1",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame1, text="Basic_Items", width="20",fg="white", bg="red", font='ariel60blue')
    label.place(x=200, y=10)
    
    label = Label(frame1, text="OF", width="20",fg="white", bg="red", font='ariel60blue')
    label.place(x=200, y=40)

    label = Label(frame1, text="First_Aid Box", width="20",fg="white", bg="red", font='ariel60blue')
    label.place(x=200, y=70)
    
    label = Label(frame2, text="Dettol",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame5, text="2",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame6, text="Cotton_Wools",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame9, text="3",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame10, text="Bandages",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame13, text="4",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame14, text="Pins",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame17, text="5",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame18, text="Thermometer",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame21, text="6", fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame22, text="Aspirin_Tablet",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame25, text="7",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame26, text="Gloves",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame29, text="8",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame30, text="Pain_Reliving Sprays",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame33, text="9",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame34, text="Pain_killer",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame37, text="10",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)

    label = Label(frame38, text="Instant_Cold Pack",fg="red", bg="white", font='ariel60blue')
    label.place(x=0, y=0)


    # Buttons
    btn = Button(frame4, text="Book", width="10", fg="red",bg="white", font='ariel10bold', command=click)
    btn.place(x=0, y=0)
    btn = Button(frame8, text="Book", width="10",fg="red", bg="white", font='ariel10bold', command=click)
    btn.place(x=0, y=0)
    btn = Button(frame12, text="Book", width="10",fg="red", bg="white", font='ariel10bold', command=click)
    btn.place(x=0, y=0)
    btn = Button(frame16, text="Book", width="10",fg="red", bg="white", font='ariel10bold', command=click)
    btn.place(x=0, y=0)
    btn = Button(frame20, text="Book", width="10",fg="red", bg="white", font='ariel10bold', command=click)
    btn.place(x=0, y=0)
    btn = Button(frame24, text="Book", width="10",fg="red", bg="white", font='ariel10bold', command=click)
    btn.place(x=0, y=0)
    btn = Button(frame28, text="Book", width="10",fg="red", bg="white", font='ariel10bold', command=click)
    btn.place(x=0, y=0)
    btn = Button(frame32, text="Book", width="10",fg="red", bg="white", font='ariel10bold', command=click)
    btn.place(x=0, y=0)
    btn = Button(frame36, text="Book", width="10",fg="red", bg="white", font='ariel10bold', command=click)
    btn.place(x=0, y=0)
    btn = Button(frame40, text="Book", width="10",fg="red", bg="white", font='ariel10bold', command=click)
    btn.place(x=0, y=0)


    aid.mainloop()

#########################find ambulance#####################
def Find_Ambulance():
        fileName='Backend_data.xlsx'
        ambulance=Toplevel()
        ambulance.title('Request For An Ambulance')
        ambulance.geometry('500x700')
        ambulance.iconbitmap("v.ico")
        ambulance.resizable(0,0)
  
        ha=Frame(ambulance,height=30,width=500,bg='red')
        ha.place(x=0,y=5)
        topic=Label(ha,text="Book Ambulance",fg="white",bg="red",font=("Helvetica"))
        topic.place(x=180,y=0)
        he=Frame(ambulance,height=290,width=396,bg="white")
        he.place(x=40,y=120)

        img100=ImageTk.PhotoImage(Image.open("ambu.png"))
        Label(he,image=img100).place(x=0,y=0)
    
        def img_click():

            img= Tk()
            img.title("Emergency-Alert")
            img.geometry("700x400+300+200")
            img.resizable(0,0)
            img.configure(bg="red")
            img_label=Label(img,text='Ambulance Booking Form',bg="red",fg="white",font=("Times New Roman", 24,'bold'))
            img_label.place(x=150,y=10)
            file=pathlib.Path(fileName)
            if file.exists():
                pass
            else:
                file=Workbook()
                sheet=file.active 
                sheet["A1"]="Username"
                # sheet["B1"]="Full name"
                sheet["B1"]="Hospital Name"
                sheet["C1"]="Phone Number"
                sheet["D1"]="Gender"
                sheet["E1"]="Address"
                file.save(fileName)
                file.close()

            def submit():
                # name=nameValue.get()
                hospitalname=hospitalnameValue.get()
                number=numberValue.get()
                gender=gender_combobox.get()
                address=addressEntry.get(1.0,END)

                file=openpyxl.load_workbook(fileName)
                sheet=file.active
                sheet.cell(column=1,row=sheet.max_row+1,value=user.username)
                # sheet.cell(column=2,row=sheet.max_row,value=name)
                sheet.cell(column=2,row=sheet.max_row,value=hospitalname)
                sheet.cell(column=3,row=sheet.max_row,value=number)
                sheet.cell(column=4,row=sheet.max_row,value=gender)
                sheet.cell(column=5,row=sheet.max_row,value=address)

                file.save(fileName)
                file.close()

                messagebox.showinfo("Ambulance Booking","Ambulance Is On Its Way")

                # nameValue.set("")
                hospitalnameValue.set("")
                numberValue.set("")
                addressEntry.delete(1.0,END)

            def clear():
            #  nameValue.set("")
             hospitalnameValue.set("")
             numberValue.set("")
             addressEntry.delete(1.0,END)


            img_frame=Frame(img,width=390,height=3,bg='white')
            img_frame.place(x=150,y=55)

            # Label(img,text="Name",font=23,bg="red",fg="white").place(x=50,y=100)
            Label(img,text="Hospital Name",font=23,bg="red",fg="white").place(x=50,y=150)
            Label(img,text="Number",font=23,bg="red",fg="white").place(x=50,y=200)
            Label(img,text="Gender",font=23,bg="red",fg="white").place(x=390,y=200)
            Label(img,text="Address",font=23,bg="red",fg="white").place(x=50,y=250)
            
            # nameValue=StringVar(img)
            hospitalnameValue=StringVar(img)
            numberValue=StringVar(img)

            # nameEntry=Entry(img,textvariable=nameValue,width=15,bd=0,font=20)
            hospitalnameEntry=Entry(img,textvariable=hospitalnameValue,width=15,bd=0,font=20)
            numberEntry=Entry(img,textvariable=numberValue,width=15,bd=0,font=20)
            addressEntry=Text(img,width=25,height=4,bd=4)

            # nameEntry.place(x=200,y=100)
            hospitalnameEntry.place(x=200,y=150)
            numberEntry.place(x=200,y=200)
            addressEntry.place(x=200,y=250)

            gender_combobox=Combobox(img,value=["Male","Female"],font="arial 14",state="r",width=7)
            gender_combobox.place(x=470,y=200)
            gender_combobox.set("")

        

            Button(img,text="Submit",bg="white",fg="red",width=15,height=2,command=submit).place(x=200,y=350)
            Button(img,text="Clear",bg="white",fg="red",width=15,height=2,command=clear).place(x=340,y=350)
            Button(img,text="Exit",bg="white",fg="red",width=15,height=2,command=lambda:img.destroy()).place(x=480,y=350)
            
            img.mainloop()

       
        def releaseAmbulance():
            os.remove(fileName)
            messagebox.showinfo(message="Ambulance has been released")
            configureButton()
        
        
        btn=None
        releaseButton=None

        def configureButton():
            nonlocal releaseButton, btn


            if(releaseButton is not None):
                releaseButton.destroy()

            if(btn is not None):
                btn.destroy()

            #####checkif ambulance is available or not    
                   
            if(os.path.exists(fileName)):
                file=load_workbook(fileName,True)
                sheet=file.active 
                data = sheet.max_row
                file.close()
                if(data<2):
                     btn=Button(ambulance,text="Available",fg="white",bg="red",font=("Helvetica"),command=img_click)
                else:
                    btn=Button(ambulance,text="Unavailable",fg="white",bg="red",font=("Helvetica"),state='disabled')
                    excel_data=pd.read_excel(fileName, 0)
                    if(user.isAdmin or ('Username' in list(excel_data.columns) and  list(excel_data['Username'])[0]==user.username)):
                        releaseButton=Button(ambulance,text="Release Ambulance",fg="white",bg="red",font=("Helvetica"),command=releaseAmbulance)
                        releaseButton.place(x=200,y=500)
                    
                    # excel_data=pd.read_excel(fileName, sheet)
                    
  
            else:
                btn=Button(ambulance,text="Available",fg="white",bg="red",font=("Helvetica"),command=img_click)
            btn.place(x=80,y=500)
            
        configureButton()


        ambulance.mainloop()

##############################Customercare#########################
def customercare():
        customer=Toplevel()
        customer.title("CUSTOMER CARE",)
        customer.geometry("880x600")
        customer.iconbitmap("")
        customer.configure(bg="red")
        customer.resizable(0,0)

        win=ttk.Treeview(customer)
        Hospital="Customer Care"

        def reverse(tuples):
            new=tuples[::-1]
            return new

        def insert(id,name,address,number,query):
            cutomersupport.addcustomer(name,address,number,query,user.username)
            messagebox.showinfo("Customer Support","Booked sucessfully")

        def delete_data_database(data):
            cutomersupport.deletecustomer(int(data))
            messagebox.showinfo("Customer Support","Deleted Sucessfully")

        def update_database(id,name,address,number,query):
            cutomersupport.updatecustomer(int(id),name,address,number,query)
            messagebox.showinfo("Customer Support","Update Sucessfully")

        def read():
            if(not user.isAdmin):
                return cutomersupport.getcustomerListForUser(user.username)
            else:
                return cutomersupport.getcustomerListForAdmin()
            

        def insert_data():

        #   id=str(entryID.get())
            name=str(entryName.get())
            address=str(entryAddress.get())
            number=str(entryNumber.get())
            query=str(entryQuery.get())

            if id==""or id==" ":
                print("Error Inserting Id")
            if name==""or name==" ":
                print("Error Inserting Name") 
            if address==""or address==" ":
                print("Error Inserting Address")
            if number==""or number==" ":
                print("Error Inserting Number") 
            if query==""or query==" ":
                print("Error Inserting Query")    

            else:
                insert(str(id),str(name),str(address),str(number),str(query))

            for data in win.get_children():
                win.delete(data)
            for result in reverse(read()):
                win.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))

                win.tag_configure("orow",background="Red",font=("Times New Roman",8))
                win.place(y=370)
            

        def delete_data():
            selected_app=win.selection()[0]
            deleteData=str(win.item(selected_app)["values"][0])
            delete_data_database(deleteData)

            for data in win.get_children():
                win.delete(data)
            for result in reverse(read()):
                win.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
            
            win.tag_configure("orow",background="Red",font=("Times New Roman",8))
            win.place(y=370) 
            
        def update_data():             
            if(len(win.selection())<1): return
            selected_app=win.selection()[0]
            updatedata=str(win.item(selected_app)["values"][0])
            update_database(updatedata,entryName.get(),entryAddress.get(),entryNumber.get(),entryQuery.get())

            for data in win.get_children():
                win.delete(data)
            for result in reverse(read()):
                win.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
            
            win.tag_configure("orow",background="Red",font=("Times New Roman",8))
            win.place(y=370)


        title=Label(customer,text="Customer Care Registration",bg='red',fg='black',font=("Times New Roman",18),bd=2)
        title.place(x=500,y=10)

        frame=Frame(customer,height=3,width=250)
        frame.place(x=510,y=40)

        img=ImageTk.PhotoImage(Image.open("customercare1.png"))
        Label(customer,image=img,width=350,bg="red").place(x=10,y=0)

        # id=Label(customer,text="ID",font=("Times New Roman",15))
        name=Label(customer,text="Name:-",font=("Times New Roman",17),bg="red",fg="black")
        address=Label(customer,text="Address:-",font=("Times New Roman",17),bg="red",fg="black")
        number=Label(customer,text="Contact Number:-",font=("Times New Roman",17),bg="red",fg="Black")
        query=Label(customer,text="Query:-",font=("Times New Roman",17),bg="red",fg="black")

        # id.grid(row=1,column=0,padx=5,pady=5)
        name.place(x=400,y=60)
        address.place(x=400,y=100)
        number.place(x=400,y=140)
        query.place(x=400,y=190)

        # entryID=Entry(customer,width=25,bd=5,font=("Times New Roman",15))
        entryName=Entry(customer,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryAddress=Entry(customer,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryNumber=Entry(customer,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)
        entryQuery=Entry(customer,width=25,bd=2,font=("Times New Roman",15),highlightbackground="white",highlightthickness=3)

        # entryID.grid(row=1,column=1,columnspan=3,padx=5,pady=5)
        entryName.place(x=600,y=60)
        entryAddress.place(x=600,y=100)
        entryNumber.place(x=600,y=140)
        entryQuery.place(x=600,y=180)


        submit=Button(customer,text="Submit",padx=1,pady=1,width=5,bd=1,font=("Times New Roman",15),bg="RED",fg="white",command=insert_data)
        submit.place(x=480,y=240)

        update=Button(customer,text="Update",padx=1,pady=1,width=5,bd=1,font=("Times New Roman",15),bg="Red",fg="white",command=update_data)
        update.place(x=580,y=240)

        delete=Button(customer,text="Delete",padx=1,pady=1,width=5,bd=1,font=("Times New Roman",15),bg="Red",fg="white",command=delete_data)
        delete.place(x=680,y=240)

        #####for treeview
        style=ttk.Style()
        style.configure("Treeview.Heading",font=("Times New Roman",15))


        win["columns"]=("ID","Name","Address","Number","Query")
        win.column("#0", width=0,stretch=NO)
        win.column("ID",anchor=W,width=100)
        win.column("Name",anchor=W,width=150)
        win.column("Address",anchor=W,width=140)
        win.column("Number",anchor=W,width=140)
        win.column("Query",anchor=W,width=180)

        # win.heading("#0", width=0,stretch=NO)
        win.heading("ID",text="ID",anchor=W)
        win.heading("Name",text="Name",anchor=W)
        win.heading("Address",text="Address",anchor=W)
        win.heading("Number",text="Number",anchor=W)
        win.heading("Query",text="Query",anchor=W)

        for data in win.get_children():
            win.delete(data)
        for result in reverse(read()):
            win.insert(parent="",index="end",iid=result[0],text="",values=(result),tag=str(result[0]))
            
            
        win.tag_configure("orow",background="Red",font=("Times New Roman",8))
        win.place(x=50,y=370) 

        customer.mainloop()



#############################Registration#######################################

def signup_command():
    root1=Toplevel(root)


    root1.title("Sign up")
    root1.geometry("925x500+300+200")
    root1.configure(bg="#fff")
    root1.resizable(0,0)
    root1.iconbitmap("v.ico")



    def signup():
        username=name.get()
        password=system.get()
        confirm_password=confirm_system.get()
        try:
            if password!=confirm_password:
                raise Exception("Password and confirm password do not match")
            database.createUser(username,password)
            messagebox.showinfo("Signup","Sucessfully sign up")
            root1.destroy()
        except BaseException as ex:
            messagebox.showerror("Error",str(ex))  


    def sign():
        root1.destroy()        
                

    abc=ImageTk.PhotoImage(Image.open("j.png"))
    Label(root1,image=abc,bg="white").place(x=15,y=0)


    frame=Frame(root1,width=345,height=390,bg="#fff")
    frame.place(x=550,y=70)


    topic=Label(frame,text="Sign up",fg="#57a1f8",bg="white",font=("Microsoft Yahei UI Light",23,"bold"))
    topic.place(x=100,y=5)



    def on_enter(e):
        name.delete(0,"end")
    def on_leave(e):
        if name.get()=="":
            name.insert(0,"Username")    



    name=Entry(frame,width=25,fg="black",border=0,bg="white",font=("Microsoft Yahei UI Light",12,"bold"))
    name.place(x=30,y=80)
    name.insert(0,"Username")
    name.bind("<FocusIn>",on_enter)
    name.bind("<FocusOut>",on_leave)


    Frame(frame,width=295,height=2,bg="black").place(x=25,y=107)


    def on_enter(e):
        system.delete(0,"end")
    def on_leave(e):
        if system.get() =="":
            system.insert(0,"Password")



    system=Entry(frame,width=25,fg="black",border=0,bg="white",font=("Microsoft Yahei UI Light",12,"bold"))
    system.place(x=30,y=150)
    system.insert(0,"Password")
    system.bind("<FocusIn>", on_enter)
    system.bind("<FocusOut>",on_leave)

    Frame(frame,width=295,height=2,bg="black").place(x=25,y=177)



    def on_enter(e):
        confirm_system.delete(0,"end")
    def on_leave(e): 
        if confirm_system.get()=="":
            confirm_system.insert(0,"Confirm Password")



    confirm_system=Entry(frame,width=25,fg="black",border=0,bg="white",font=("Microsoft Yahei UI Light",12,"bold"))
    confirm_system.place(x=30,y=220)
    confirm_system.insert(0,"Confirm Password")
    confirm_system.bind("<FocusIn>", on_enter)
    confirm_system.bind("<FocusOut>",on_leave)

    Frame(frame,width=295,height=2,bg="black").place(x=25,y=247)



    Button(frame,width=39,pady=7,text="Sign up",bg="#57a1f8",fg="white",border=0,command=signup).place(x=35,y=280)
    set=Label(frame,text="I have an account.",fg="black",bg="white",font=("Microsoft Yahei UI Light",9))
    set.place(x=90,y=340)

    signin=Button(frame,width=6,text="Sign in",border=0,bg="white",cursor="hand2",fg="#57a1f8",command=sign)
    signin.place(x=200,y=340)


    root1.mainloop()
   


 ##########################################################################  



frame=Frame(root,width=345,height=345,bg="white")
frame.place(x=475,y=70)
topic=Label(frame,text="Sign in",fg="#57a1f8",bg="white",font=("Microsoft YaHei UI Light",23,"bold"))
topic.place(x=110,y=5)


def on_enter(e):
    name.delete(0,"end")
def on_leave(e):
    username=name.get() 
    if name =="":
        name.insert(0,"Username")


name=Entry(frame,width=25,fg="black",border=0,bg="white",font=("Microsoft YaHei UI Light",11,"bold"))
name.place(x=30,y=80)
name.insert(0,"Username")
name.bind("<FocusIn>", on_enter)
name.bind("<FocusOut>", on_leave)
Frame(frame,width=295,height=2,bg="black").place(x=25,y=107)


def on_enter(e):
    system.delete(0,"end")
def on_leave(e):
    password=system.get() 
    if password =="":
        system.insert(0,"Password")

      
system=Entry(frame,width=25,fg="black",border=0,bg="white",font=("Microsoft YaHei UI Light",11,"bold"))
system.place(x=30,y=150)
system.insert(0,"Password")
system.bind("<FocusIn>", on_enter)
system.bind("<FocusOut>", on_leave)
Frame(frame,width=295,height=2,bg="black").place(x=25,y=177)

Button(frame,width=39,pady=7,text="Log in",bg="#57a1f8",fg="white",border=0,command=signin).place(x=35,y=204)
set=Label(frame,text="Don't have an account?",fg="black",bg="white",font=("bold",9))
set.place(x=75,y=270)
sign_up=Button(frame,width=6,text="Sign Up",border=0,bg="white",cursor="hand2",fg="#57a1f8",command=signup_command)
sign_up.place(x=215,y=270)




root.mainloop()